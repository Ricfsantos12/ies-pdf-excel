#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extrator PDF → Excel por template (áreas fixas) com:
- Concatenação de sub-tabelas (não corta linhas a meio)
- Conversão numérica PT (remove pontos de milhar; vírgula como decimal)
- Formatação Excel sem separador de milhar (por defeito ou por coluna)
- Modos de saída: overwrite | increment | append
- Tolerâncias do Camelot (stream): row_tol, column_tol, shift_text, split_text

Uso:
  python pdf_para_excel_template.py <ficheiro.pdf> <template.yml> <saida.xlsx> [overwrite|increment|append]

YAML (por tabela):
  - name, page ou find_page_by, flavor (lattice|stream)
  - find_page_by: {text_contains|any_of|all_of, occurrence?}  # localizar página por texto
  - area_tlbr: [Top, Left, Bottom, Right]     # estilo Tabula (origem topo)
    # ou area: [x1, y1, x2, y2]               # estilo Camelot (origem baixo)
  - columns: [x1, x2, ...]                    # para stream (uma divisória por coluna)
  - min_rows, min_cols
  - line_scale (opcional; só lattice)
  - numeric_cols: "auto" | "all" | [lista de nomes]
  - excel_number_format: "0.############" (default interno; sem sep. de milhar)
  - excel_formats: {"Coluna X": "0", "Percentagem": "0.00%"}
  - row_tol, column_tol, shift_text, split_text (opcionais; stream)
  - concat_tables: true|false  (default true) → junta sub-tabelas verticalmente
"""

from __future__ import annotations

import sys
import re
import unicodedata
from collections import Counter
from pathlib import Path
from typing import List, Tuple, Optional, Dict

import yaml
import pandas as pd
import camelot
from pypdf import PdfReader
from openpyxl.utils import get_column_letter

INVALID_SHEET_CHARS = set(r':\/?*[]')

# ---------- helpers de nomes/paths ----------

def sanitize_sheet_name(name: str, used: Dict[str, int]) -> str:
    cleaned = "".join(ch for ch in name if ch not in INVALID_SHEET_CHARS).strip() or "Sheet"
    if len(cleaned) > 31:
        cleaned = cleaned[:31]
    base = cleaned
    if base not in used:
        used[base] = 1
        return base
    i = used[base]
    while True:
        i += 1
        candidate = (base[:28] + f"_{i}") if len(base) > 28 else f"{base}_{i}"
        if candidate not in used:
            used[base] = i
            used[candidate] = 1
            return candidate

def next_incremented_path(p: Path) -> Path:
    if not p.exists():
        return p
    stem, suf = p.stem, p.suffix or ".xlsx"
    i = 1
    while True:
        cand = p.with_name(f"{stem}_{i}{suf}")
        if not cand.exists():
            return cand
        i += 1

# ---------- utilidades PDF/Tabela ----------

def fmt_area(area: List[float]) -> str:
    return ",".join(f"{v:.3f}" for v in area)

def parse_number_pt(x):
    """Converte '1.234,56' → 1234.56. Remove pontos de milhar (inclui NBSP) e usa vírgula como decimal."""
    if x is None:
        return x
    s = str(x).strip()
    if s == "" or s.lower() in {"na", "n/a"} or s in {"-", "–", "—"}:
        return None
    # normalizações (NBSP U+00A0, thin NBSP U+202F, menos matemático)
    s = (s.replace("\u00A0", "")
           .replace("\u202F", "")
           .replace(" ", "")
           .replace("–", "-")
           .replace("—", "-")
           .replace("−", "-"))
    # remover pontos de milhar: ponto entre grupos de 3 dígitos
    s = re.sub(r"(?<=\d)[.](?=\d{3}(\D|$))", "", s)
    # vírgula decimal → ponto
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return x  # mantém como texto se não for número claro

def clean_dataframe(df: pd.DataFrame, numeric_cols="auto", promote_header: bool = True) -> pd.DataFrame:
    """
    Trim, cabeçalho (opcional) e conversão numérica PT (com fallback por célula).

    promote_header = True  → promove 1.ª linha a header se fizer sentido
    promote_header = False → mantém 1.ª linha como dados (usado em 'DR')
    """
    # trim universal (sem remover espaços internos)
    if hasattr(df, "map"):  # pandas ≥2.3
        df = df.map(lambda v: str(v).strip())
    else:
        df = df.applymap(lambda v: str(v).strip())

    # promover 1.ª linha a header, se fizer sentido e se estiver activo
    if promote_header and df.shape[0] > 0 and all(str(c).strip() != "" for c in df.iloc[0].tolist()):
        df.columns = df.iloc[0].tolist()
        df = df.iloc[1:].reset_index(drop=True)

    # heurística de “parece numérico”
    num_like = re.compile(r"^-?\s*\d[\d.,\u00A0\u202F]*\s*$")

    # escolher colunas a tratar
    if numeric_cols == "all":
        cand_cols = list(df.columns)
    elif numeric_cols == "auto":
        cand_cols = [col for col in df.columns if df[col].astype(str).str.match(num_like).mean() >= 0.5]
    elif isinstance(numeric_cols, list):
        cand_cols = [col for col in numeric_cols if col in df.columns]
    else:
        cand_cols = []

    # conversão por coluna
    for col in cand_cols:
        df[col] = df[col].apply(parse_number_pt)

    # fallback: em auto/all, tentar células isoladas que “parecem número”
    if numeric_cols in ("all", "auto"):
        for col in df.columns:
            if col in cand_cols:
                continue
            mask = df[col].astype(str).str.match(num_like)
            if mask.any():
                df.loc[mask, col] = df.loc[mask, col].apply(parse_number_pt)

    return df

def page_size(pdf_path: Path, page: int) -> Tuple[float, float]:
    r = PdfReader(str(pdf_path))
    mb = r.pages[page - 1].mediabox
    return float(mb.width), float(mb.height)

def tlbr_to_camelot(top: float, left: float, bottom: float, right: float, H: float) -> List[float]:
    x1 = left; x2 = right
    y1 = H - bottom
    y2 = H - top
    return [x1, y1, x2, y2]

def normalize_area(a: List[float]) -> List[float]:
    x1, y1, x2, y2 = a
    if x1 > x2: x1, x2 = x2, x1
    if y1 > y2: y1, y2 = y2, y1
    return [x1, y1, x2, y2]

def pad_area(a: List[float], pad: float, W: float, H: float) -> List[float]:
    x1, y1, x2, y2 = a
    return [max(0, x1 - pad), max(0, y1 - pad), min(W, x2 + pad), min(H, y2 + pad)]

def extract_table_once(pdf_path: Path, page: int, flavor: str, area: List[float],
                       columns: Optional[List[float]], line_scale: int, extra: Optional[Dict]=None):
    """Lê uma tabela Camelot com parâmetros fornecidos (inclui tolerâncias 'extra')."""
    kwargs = {
        "pages": str(page),
        "flavor": flavor,
        "table_areas": [fmt_area(area)],
        # NÃO remover espaços internos: sem o espaço aqui!
        "strip_text": "\n\r\t",
    }
    if flavor == "lattice":
        kwargs["line_scale"] = line_scale
    if flavor == "stream" and columns:
        kwargs["columns"] = [",".join(str(c) for c in columns)]
    # tolerâncias/flags opcionais vindas do YAML
    if extra:
        if "row_tol" in extra:       kwargs["row_tol"] = int(extra["row_tol"])
        if "column_tol" in extra:    kwargs["column_tol"] = int(extra["column_tol"])
        if "shift_text" in extra:    kwargs["shift_text"] = bool(extra["shift_text"])
        if "split_text" in extra:    kwargs["split_text"] = bool(extra["split_text"])
    return camelot.read_pdf(str(pdf_path), **kwargs)

def pick_first_valid(tables, min_rows: int, min_cols: int):
    for t in tables:
        df = t.df
        if df.shape[0] >= min_rows and df.shape[1] >= min_cols:
            return df
    return None

def concat_valid_tables(tables, min_rows: int, min_cols: int):
    """
    Junta verticalmente todas as sub-tabelas que passam os mínimos.
    Mantém apenas o nº de colunas mais comum e remove cabeçalhos repetidos.
    """
    candidates = [t.df for t in tables if t.df.shape[0] >= min_rows and t.df.shape[1] >= min_cols]
    if not candidates:
        return None
    col_counts = Counter(df.shape[1] for df in candidates)
    common_cols = col_counts.most_common(1)[0][0]
    parts = [df for df in candidates if df.shape[1] == common_cols]
    if not parts:
        return None
    big = parts[0].copy()
    for df in parts[1:]:
        df_part = df.iloc[1:] if df.shape[0] > 0 else df  # corta header repetido
        big = pd.concat([big, df_part], ignore_index=True)
    return big


def normalize_text(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = "".join(ch if ch.isalnum() or ch.isspace() else " " for ch in s)
    s = s.lower()
    s = " ".join(s.split())
    return s


def compact_text(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return "".join(ch.lower() for ch in s if ch.isalnum())


def extract_page_text(reader: PdfReader, page_number: int) -> str:
    try:
        txt = reader.pages[page_number - 1].extract_text() or ""
        return txt
    except Exception:
        return ""


def contains_fragment(page_text: str, fragment: str) -> bool:
    page_norm = normalize_text(page_text)
    frag_norm = normalize_text(fragment)
    if frag_norm and frag_norm in page_norm:
        return True

    page_compact = compact_text(page_text)
    frag_compact = compact_text(fragment)
    return bool(frag_compact) and frag_compact in page_compact


def page_matches(text: str, criteria: dict) -> bool:
    if "text_contains" in criteria:
        return contains_fragment(text, criteria["text_contains"])

    if "any_of" in criteria:
        return any(contains_fragment(text, x) for x in criteria["any_of"])

    if "all_of" in criteria:
        return all(contains_fragment(text, x) for x in criteria["all_of"])

    raise ValueError("find_page_by tem de incluir 'text_contains', 'any_of' ou 'all_of'")


def find_page_by_content(pdf_path: Path, criteria: dict) -> int:
    reader = PdfReader(str(pdf_path))
    occurrence = int(criteria.get("occurrence", 1))
    found = 0

    for page_number in range(1, len(reader.pages) + 1):
        text = extract_page_text(reader, page_number)
        if page_matches(text, criteria):
            found += 1
            if found == occurrence:
                return page_number

    raise RuntimeError(f"Nenhuma página encontrada para o critério: {criteria}")


def resolve_page(pdf_path: Path, tbl: dict) -> int:
    if "page" in tbl:
        return int(tbl["page"])

    if "find_page_by" in tbl:
        return find_page_by_content(pdf_path, tbl["find_page_by"])

    raise KeyError(f"Entrada '{tbl.get('name', 'Tabela')}' sem 'page' nem 'find_page_by'.")

# ---------- núcleo ----------

def main(pdf_path: str, yaml_path: str, xlsx_path: str, mode: str = "overwrite"):
    pdf_path = Path(pdf_path)
    cfg = yaml.safe_load(Path(yaml_path).read_text(encoding="utf-8"))

    if "tables" not in cfg or not isinstance(cfg["tables"], list) or len(cfg["tables"]) == 0:
        raise ValueError("O YAML deve conter uma lista 'tables' com pelo menos uma entrada.")

    mode_norm = (mode or "overwrite").strip().lower()
    if mode_norm in {"w", "write", "new"}: mode_norm = "overwrite"
    if mode_norm in {"inc", "incr"}:       mode_norm = "increment"
    if mode_norm not in {"overwrite", "increment", "append"}:
        raise ValueError("Modo inválido. Use: overwrite | increment | append")

    out_path = Path(xlsx_path)
    writer_kwargs = {"engine": "openpyxl"}

    if mode_norm == "append":
        writer_kwargs.update({"mode": "a", "if_sheet_exists": "replace"})
        final_out = out_path
    elif mode_norm == "increment":
        final_out = next_incremented_path(out_path)
        writer_kwargs.update({"mode": "w"})
    else:  # overwrite
        final_out = out_path
        writer_kwargs.update({"mode": "w"})

    # abrir ExcelWriter (se bloqueado e 'increment', tenta nome seguinte)
    try:
        writer = pd.ExcelWriter(final_out, **writer_kwargs)
    except PermissionError:
        if mode_norm in {"overwrite", "append"}:
            raise
        final_out = next_incremented_path(final_out)
        writer = pd.ExcelWriter(final_out, **writer_kwargs)

    used_sheet_names: Dict[str, int] = {}

    for tbl in cfg["tables"]:
        raw_name = str(tbl.get("name", "Tabela"))
        sheet_name = sanitize_sheet_name(raw_name, used_sheet_names)

        page = resolve_page(pdf_path, tbl)
        flavor = str(tbl.get("flavor", "lattice")).lower()
        if flavor not in {"lattice", "stream"}:
            raise ValueError(f"flavor inválido em '{raw_name}': {flavor}")

        min_rows = int(tbl.get("min_rows", 1))
        min_cols = int(tbl.get("min_cols", 1))
        numeric_cols = tbl.get("numeric_cols", "auto")
        line_scale = int(tbl.get("line_scale", 80))
        concat_opt = bool(tbl.get("concat_tables", True))

        # tolerâncias/flags (opcionais)
        extra = {}
        for key in ("row_tol", "column_tol", "shift_text", "split_text"):
            if key in tbl:
                extra[key] = tbl[key]

        # formatação Excel (default: sem separador de milhar)
        excel_default_fmt = tbl.get("excel_number_format") or "0.############"
        excel_formats = tbl.get("excel_formats", {})

        # --- calcular área Camelot ---
        W, H = page_size(pdf_path, page)
        if "area_tlbr" in tbl:
            top, left, bottom, right = map(float, tbl["area_tlbr"])
            area0 = tlbr_to_camelot(top, left, bottom, right, H)
        elif "area" in tbl:
            area0 = list(map(float, tbl["area"]))
        else:
            raise KeyError(f"Entrada '{raw_name}' sem 'area_tlbr' nem 'area'.")

        area0 = normalize_area(area0)
        columns = [float(x) for x in tbl.get("columns", [])] if tbl.get("columns") else None

        # --- tentativas com acolchoamento (mesmo flavor) ---
        tried = []
        df_clean = None
        area_used = None

        for pad in (0.0, 2.0, 5.0):
            area_try = pad_area(area0, pad, W, H) if pad else area0

            # SANITY CHECK às columns para caberem na área e virem ordenadas
            cols = columns[:] if columns else None
            if cols:
                x1, y1, x2, y2 = area_try
                eps = 0.5  # empurra 0.5 pt para dentro, se estiver no limite
                cols_sorted = sorted(set(float(c) for c in cols))
                cols_clamped = [min(x2 - eps, max(x1 + eps, c)) for c in cols_sorted]
                if cols_clamped != cols:
                    print(f"[WARN] Ajustei columns de {cols} -> {cols_clamped} para caber na área {area_try}")
                cols = cols_clamped

            try:
                tables = extract_table_once(pdf_path, page, flavor, area_try, cols, line_scale, extra=extra)
            except Exception as e:
                tried.append((area_try, f"ERRO {e.__class__.__name__}: {e}"))
                continue

            if len(tables) == 0:
                tried.append((area_try, "SEM TABELA"))
                continue

            # juntar partes (ou escolher a 1.ª válida)
            if concat_opt:
                df_raw = concat_valid_tables(tables, min_rows, min_cols)
            else:
                df_raw = pick_first_valid(tables, min_rows, min_cols)

            if df_raw is not None:
                # Para a folha "DR": NÃO promover 1.ª linha a cabeçalho
                promote_header = not (sheet_name.upper() == "DR")
                df_clean = clean_dataframe(df_raw, numeric_cols=numeric_cols, promote_header=promote_header)
                area_used = area_try
                break
            else:
                tried.append((area_try, f"sem partes válidas (min_rows={min_rows}, min_cols={min_cols})"))

        if df_clean is None:
            detalhes = "; ".join([f"{[round(v,3) for v in a]} -> {msg}" for a, msg in tried])
            raise RuntimeError(
                f"Nenhuma extração válida para '{raw_name}' (pág. {page}) com flavor='{flavor}'. "
                f"Dica: em 'stream', define 'columns' (todas as divisórias) e ajusta 'area_tlbr' 1–2 pt para dentro da grelha. "
                f"Tentativas: {detalhes}"
            )

        # --- escrever folha ---
        is_dr = (sheet_name.upper() == "DR")
        header_flag = not is_dr  # em DR não escreve linha de cabeçalho

        df_clean.to_excel(writer, sheet_name=sheet_name, index=False, header=header_flag)
        ws = writer.sheets[sheet_name]

        # larguras auto
        for i, col in enumerate(df_clean.columns, start=1):
            try:
                max_len = max([len(str(col))] + [len(str(v)) for v in df_clean[col].tolist()])
                ws.column_dimensions[get_column_letter(i)].width = min(max(12, max_len + 2), 60)
            except Exception:
                pass

        # aplicar formatos numéricos (sem separador de milhar por defeito)
        n_rows = df_clean.shape[0]
        # se há cabeçalho, dados começam na linha 2; se não, na linha 1
        row_start = 2 if header_flag else 1

        for col_idx, col_name in enumerate(df_clean.columns, start=1):
            fmt = excel_formats.get(str(col_name), excel_default_fmt)
            for row_idx in range(row_start, n_rows + row_start):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = fmt

        print(f"[INFO] {raw_name} → '{sheet_name}': {flavor}, area={area_used}, shape={df_clean.shape}")

    writer.close()
    print(f"✔ Exportado para {final_out}")

# ---------- CLI ----------

if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Uso: python pdf_para_excel_template.py <ficheiro.pdf> <template.yml> <saida.xlsx> [overwrite|increment|append]")
        sys.exit(1)
    pdf = sys.argv[1]
    yml = sys.argv[2]
    xlsx = sys.argv[3]
    mode = sys.argv[4] if len(sys.argv) >= 5 else "overwrite"
    main(pdf, yml, xlsx, mode)


